"""
Celery tasks for employee CSV/XLSX import and export.
"""
import csv
import io
import os
from datetime import date
from decimal import Decimal, InvalidOperation

from celery import shared_task
from django.core.files.base import ContentFile

from core.models import BackgroundTask, Country, Currency, Department, JobTitle

from .models import Employee, SalaryRecord
from .services import convert_to_usd


IMPORT_REQUIRED_COLUMNS = {
    "full_name",
    "department",
    "job_title",
    "country",
    "base_salary",
    "effective_date",
    "employment_type",
}


def _rows_from_csv(file_obj):
    """Yield dicts from a CSV file object."""
    content = file_obj.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    headers = set(reader.fieldnames or [])
    return headers, reader


def _rows_from_xlsx(file_obj):
    """Yield dicts from the first sheet of an XLSX workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return set(), iter([])

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    header_set = set(headers)

    def row_generator():
        for row in rows_iter:
            # Skip completely empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_dict = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                # Convert non-string values (numbers, dates) to strings
                row_dict[h] = str(val).strip() if val is not None else ""
            yield row_dict

    return header_set, row_generator()


def _process_row(row, row_num, departments, countries, job_titles, valid_emp_types):
    """
    Validate and process a single import row.
    Returns (employee, record) on success, or (None, error_dict) on failure.
    """
    row_errors = []

    # Required field checks
    full_name = (row.get("full_name") or "").strip()
    if not full_name:
        row_errors.append("full_name is required.")

    # Department lookup
    dept_name = (row.get("department") or "").strip().lower()
    dept = departments.get(dept_name)
    if not dept:
        row_errors.append(f"Department '{row.get('department', '')}' not found.")

    # Country lookup
    country_name = (row.get("country") or "").strip().lower()
    country = countries.get(country_name)
    if not country:
        row_errors.append(f"Country '{row.get('country', '')}' not found.")

    # Job title lookup (requires dept)
    jt_title = (row.get("job_title") or "").strip().lower()
    job_title = None
    if dept:
        job_title = job_titles.get((jt_title, dept.id))
        if not job_title:
            row_errors.append(
                f"Job title '{row.get('job_title', '')}' not found in department '{dept.name}'."
            )

    # Employment type
    emp_type = (row.get("employment_type") or "").strip().lower()
    if emp_type not in valid_emp_types:
        row_errors.append(
            f"Invalid employment_type '{row.get('employment_type', '')}'. "
            f"Must be one of: {', '.join(sorted(valid_emp_types))}."
        )

    # Base salary
    try:
        base_salary = Decimal(row.get("base_salary", "").strip())
        if base_salary <= 0:
            row_errors.append("base_salary must be > 0.")
    except (InvalidOperation, ValueError):
        row_errors.append(f"Invalid base_salary: '{row.get('base_salary', '')}'.")
        base_salary = None

    # Effective date
    try:
        raw_date = (row.get("effective_date") or "").strip()
        # Handle date objects from XLSX (openpyxl may return datetime)
        effective_date = date.fromisoformat(raw_date[:10]) if raw_date else None
        if effective_date is None:
            row_errors.append("effective_date is required.")
    except (ValueError, TypeError):
        row_errors.append(
            f"Invalid effective_date: '{row.get('effective_date', '')}'. Use YYYY-MM-DD."
        )
        effective_date = None

    if row_errors:
        return None, {"row": row_num, "errors": row_errors}

    # Create employee + salary record
    employee = Employee.objects.create(
        full_name=full_name,
        department=dept,
        job_title=job_title,
        country=country,
        local_currency=country.default_currency,
        employment_type=emp_type,
    )

    record = SalaryRecord(
        employee=employee,
        base_salary=base_salary,
        effective_date=effective_date,
        variable_bonus_pct=Decimal(row.get("variable_bonus_pct", "0") or "0"),
        source="csv_import",
        hr_note="Imported via file upload.",
    )
    record.save()

    Employee.objects.filter(pk=employee.pk).update(current_salary_record=record)

    return employee, None


@shared_task(bind=True)
def process_csv_import(self, task_id):
    """
    Parse an uploaded CSV or XLSX file, validate each row, create
    Employee + SalaryRecord for valid rows, and report per-row results.
    """
    task = BackgroundTask.objects.get(id=task_id)
    task.status = "processing"
    task.save(update_fields=["status"])

    results = {"rows_imported": 0, "rows_failed": 0, "errors": []}

    try:
        file_name = task.file.name or ""
        ext = os.path.splitext(file_name)[1].lower()

        # Parse based on file extension
        if ext == ".xlsx":
            task.file.open("rb")
            headers, row_iter = _rows_from_xlsx(task.file)
        elif ext in (".csv", ""):
            task.file.open("rb")
            headers, row_iter = _rows_from_csv(task.file)
        else:
            task.status = "failed"
            task.result_data = {"error": f"Unsupported file format: '{ext}'. Use .csv or .xlsx."}
            task.save(update_fields=["status", "result_data"])
            return

        # Validate headers
        missing = IMPORT_REQUIRED_COLUMNS - headers
        if missing:
            task.status = "failed"
            task.result_data = {"error": f"Missing columns: {', '.join(sorted(missing))}"}
            task.save(update_fields=["status", "result_data"])
            return

        # Cache lookups
        departments = {d.name.lower(): d for d in Department.objects.all()}
        countries = {c.name.lower(): c for c in Country.objects.select_related("default_currency").all()}
        job_titles = {(jt.title.lower(), jt.department_id): jt for jt in JobTitle.objects.all()}

        valid_emp_types = {"full_time", "part_time", "contractor"}

        for row_num, row in enumerate(row_iter, start=2):
            employee, error = _process_row(
                row, row_num, departments, countries, job_titles, valid_emp_types
            )
            if error:
                results["rows_failed"] += 1
                results["errors"].append(error)
            else:
                results["rows_imported"] += 1

        task.status = "completed"
        task.result_data = results
        task.save(update_fields=["status", "result_data"])

    except Exception as exc:
        task.status = "failed"
        task.result_data = {"error": str(exc)}
        task.save(update_fields=["status", "result_data"])
        raise


@shared_task(bind=True)
def generate_export_file(self, task_id, filters=None, export_format="csv"):
    """
    Build a CSV or XLSX from filtered employee queryset with USD conversions.
    """
    task = BackgroundTask.objects.get(id=task_id)
    task.status = "processing"
    task.save(update_fields=["status"])

    try:
        qs = Employee.objects.select_related(
            "department", "job_title", "country", "local_currency", "current_salary_record"
        ).filter(is_active=True)

        if filters:
            if "department" in filters:
                qs = qs.filter(department_id=filters["department"])
            if "country" in filters:
                qs = qs.filter(country_id=filters["country"])
            if "employment_type" in filters:
                qs = qs.filter(employment_type=filters["employment_type"])
            if "is_active" in filters:
                qs = qs.filter(is_active=filters["is_active"])

        headers = [
            "employee_code",
            "full_name",
            "department",
            "job_title",
            "country",
            "currency",
            "employment_type",
            "base_salary",
            "base_salary_usd",
            "variable_bonus_pct",
            "effective_date",
        ]

        rows = []
        for emp in qs:
            salary = emp.current_salary_record
            usd = convert_to_usd(
                salary.base_salary if salary else None,
                emp.local_currency,
            )
            rows.append([
                emp.employee_code,
                emp.full_name,
                emp.department.name,
                emp.job_title.title,
                emp.country.name,
                emp.local_currency.code,
                emp.employment_type,
                str(salary.base_salary) if salary else "",
                str(usd["usd_value"]) if usd["usd_value"] else "",
                str(salary.variable_bonus_pct) if salary else "",
                str(salary.effective_date) if salary else "",
            ])

        if export_format == "xlsx":
            _write_xlsx(task, headers, rows)
        else:
            _write_csv(task, headers, rows)

        task.status = "completed"
        task.save(update_fields=["status"])

    except Exception as exc:
        task.status = "failed"
        task.result_data = {"error": str(exc)}
        task.save(update_fields=["status", "result_data"])
        raise


def _write_csv(task, headers, rows):
    """Write CSV content to task file."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    task.file.save(
        f"employee_export_{task.id}.csv",
        ContentFile(output.getvalue().encode("utf-8")),
    )


def _write_xlsx(task, headers, rows):
    """Write XLSX content to task file."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(headers)
    for row in rows:
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    task.file.save(f"employee_export_{task.id}.xlsx", ContentFile(output.read()))
